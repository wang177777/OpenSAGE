"""Generate release manifests and enforce the public-release quality gate."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import logging
import re
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader

logging.getLogger("pypdf").setLevel(logging.ERROR)


ROOT = Path(__file__).resolve().parents[1]
TABLES = ROOT / "data/publication_tables"
MANIFESTS = ROOT / "data/manifests"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate the OpenSAGE public release")
    parser.add_argument(
        "--check",
        action="store_true",
        help="Fail if regenerated manifests differ from committed manifests.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str], check: bool) -> None:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    content = buffer.getvalue()
    if check:
        if not path.exists() or path.read_text(encoding="utf-8") != content:
            raise AssertionError(f"Committed manifest is stale: {path.relative_to(ROOT)}")
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")


def validate_workbooks() -> None:
    expected = {
        "Supplementary_Data_1.xlsx": {"T001_56_Disease_List": 57, "T002_Matched_80": 81, "S02": 4},
        "Supplementary_Data_2.xlsx": {"S12A": 4, "S12B": 4},
    }
    for filename, sheets in expected.items():
        path = ROOT / "data/supplementary" / filename
        workbook = load_workbook(path, data_only=True)
        if workbook.properties.lastModifiedBy:
            raise AssertionError(f"Workbook exposes lastModifiedBy: {filename}")
        for sheet, rows in sheets.items():
            if sheet not in workbook.sheetnames or workbook[sheet].max_row != rows:
                raise AssertionError(f"Unexpected workbook sheet/row count: {filename}:{sheet}")
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.comment is not None:
                        raise AssertionError(f"Workbook comment remains: {filename}:{sheet.title}!{cell.coordinate}")
        for csv_path in TABLES.glob("*.csv"):
            if csv_path.stem not in workbook.sheetnames:
                continue
            with csv_path.open(encoding="utf-8-sig", newline="") as handle:
                expected_rows = list(csv.reader(handle))
            sheet = workbook[csv_path.stem]
            actual_rows = [
                ["" if cell.value is None else str(cell.value) for cell in row]
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=len(expected_rows),
                    min_col=1,
                    max_col=max(len(row) for row in expected_rows),
                )
            ]
            normalized_expected = [
                row + [""] * (len(actual_rows[0]) - len(row)) for row in expected_rows
            ]
            if actual_rows != normalized_expected:
                raise AssertionError(
                    f"Workbook/CSV mismatch: {filename}:{csv_path.stem}"
                )


def secret_and_path_scan() -> None:
    patterns = {
        "OpenAI-style key": re.compile(rb"\bsk-[A-Za-z0-9_-]{16,}\b"),
        "Tavily key": re.compile(rb"\btvly-[A-Za-z0-9_-]{16,}\b"),
        "LangSmith key": re.compile(rb"\blsv2_[A-Za-z0-9_-]{16,}\b"),
        "GitHub token": re.compile(rb"\b(?:ghp|github_pat)_[A-Za-z0-9_]{16,}\b"),
        "Private key": re.compile(rb"-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----"),
        "Local user path": re.compile(rb"(?:/Users/|[A-Za-z]:\\\\Users\\\\)"),
    }
    def scan_payload(source: str, data: bytes) -> None:
        for name, pattern in patterns.items():
            if pattern.search(data):
                raise AssertionError(f"{name} detected in {source}")

    structured_suffixes = {".pdf", ".xlsx", ".png", ".jpg", ".jpeg", ".tif", ".tiff"}
    ignored_binary_suffixes = {".woff", ".woff2"}
    for path in ROOT.rglob("*"):
        if (
            not path.is_file()
            or any(
                part in {".git", "__pycache__", "node_modules", ".venv", "generated", "dist"}
                for part in path.parts
            )
            or path.suffix.lower() in structured_suffixes | ignored_binary_suffixes
        ):
            continue
        if path.resolve() == Path(__file__).resolve():
            continue  # This file necessarily contains the detection patterns themselves.
        scan_payload(path.relative_to(ROOT).as_posix(), path.read_bytes())

    for path in ROOT.rglob("*.pdf"):
        if any(part in {".git", "generated"} for part in path.parts):
            continue
        reader = PdfReader(path)
        metadata = "\n".join(f"{key}: {value}" for key, value in (reader.metadata or {}).items())
        text = metadata + "\n" + "\n".join((page.extract_text() or "") for page in reader.pages)
        scan_payload(path.relative_to(ROOT).as_posix(), text.encode("utf-8", errors="ignore"))

    for path in ROOT.rglob("*.xlsx"):
        if ".git" in path.parts:
            continue
        workbook = load_workbook(path, data_only=False, read_only=True)
        values = [
            str(value)
            for value in (
                workbook.properties.title,
                workbook.properties.subject,
                workbook.properties.creator,
                workbook.properties.lastModifiedBy,
                workbook.properties.description,
                workbook.properties.keywords,
            )
            if value
        ]
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                values.extend(str(cell.value) for cell in row if cell.value is not None)
        scan_payload(path.relative_to(ROOT).as_posix(), "\n".join(values).encode("utf-8", errors="ignore"))

    for suffix in ("*.png", "*.jpg", "*.jpeg", "*.tif", "*.tiff"):
        for path in ROOT.rglob(suffix):
            if any(part in {".git", "generated"} for part in path.parts):
                continue
            with Image.open(path) as image:
                metadata = [f"{key}: {value}" for key, value in image.info.items()]
                exif = image.getexif()
                metadata.extend(f"{key}: {value}" for key, value in exif.items())
            scan_payload(path.relative_to(ROOT).as_posix(), "\n".join(metadata).encode("utf-8", errors="ignore"))


def release_hashes(check: bool) -> None:
    excluded = {
        MANIFESTS / "release_sha256.csv",
        ROOT / "results/reproduced_summary.json",
    }
    rows = []
    for path in sorted(ROOT.rglob("*")):
        if (
            not path.is_file()
            or any(
                part in {".git", "__pycache__", "node_modules", ".venv", "dist"}
                for part in path.parts
            )
            or path in excluded
        ):
            continue
        if "generated" in path.parts or path.name == ".DS_Store":
            continue
        rows.append({"file_path": path.relative_to(ROOT).as_posix(), "sha256": sha256(path)})
    write_csv(MANIFESTS / "release_sha256.csv", rows, ["file_path", "sha256"], check)


def main() -> None:
    args = parse_args()
    MANIFESTS.mkdir(parents=True, exist_ok=True)
    validate_workbooks()
    secret_and_path_scan()
    release_hashes(args.check)
    print("PASS - public tables, workbooks, figures, code, and release hashes")
    print("PASS - text, PDF, workbook, image-metadata, credential, and local-path checks")


if __name__ == "__main__":
    main()
